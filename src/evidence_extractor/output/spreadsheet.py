import logging
import re

from openpyxl import Workbook

from evidence_extractor.models.schemas import ArticleExtraction

logger = logging.getLogger(__name__)


def _clean_sheet_title(title: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "", title)[:31]


def export_to_excel(extraction: ArticleExtraction, output_path: str):
    logger.info(f"Starting export of findings to Excel file: '{output_path}'")
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Field", "Value", "Validation Status"])
    ws_summary.append(["Source Filename", extraction.source_filename, "N/A"])
    ws_summary.append(["Extracted Title", extraction.title, "N/A"])
    ws_summary.append(["Generated Summary", extraction.summary, "N/A"])
    if extraction.pico_elements:
        pico = extraction.pico_elements
        ws_summary.append(
            [
                "PICO - Population",
                pico.population,
                pico.correction_metadata.status.value,
            ]
        )
        ws_summary.append(
            [
                "PICO - Intervention",
                pico.intervention,
                pico.correction_metadata.status.value,
            ]
        )
        ws_summary.append(
            [
                "PICO - Comparison",
                pico.comparison,
                pico.correction_metadata.status.value,
            ]
        )
        ws_summary.append(
            ["PICO - Outcome", pico.outcome, pico.correction_metadata.status.value]
        )
    for score in extraction.quality_scores:
        ws_summary.append(
            [
                f"Quality Score - {score.score_name}",
                score.score_value,
                score.correction_metadata.status.value,
            ]
        )
        ws_summary.append(
            [
                "Quality Justification",
                score.justification,
                score.correction_metadata.status.value,
            ]
        )
    if extraction.claims:
        ws_claims = wb.create_sheet("Claims")
        ws_claims.append(
            [
                "Claim Text",
                "Uncertainty Annotation",
                "Page Number",
                "Validation Status",
                "Reviewer Comment",
            ]
        )
        for claim in extraction.claims:
            ws_claims.append(
                [
                    claim.claim_text,
                    claim.uncertainty_annotation,
                    claim.provenance.page_number,
                    claim.correction_metadata.status.value,
                    claim.correction_metadata.reviewer_comment,
                ]
            )
    for i, table in enumerate(extraction.tables):
        sheet_title = _clean_sheet_title(
            f"Table {i + 1} (Page {table.provenance.page_number})"
        )
        ws_table = wb.create_sheet(sheet_title)
        if table.summary:
            ws_table.append([f"AI Summary: {table.summary}"])
            ws_table.append([])

        if table.table_data:
            headers = list(table.table_data[0].keys())
            ws_table.append(headers)
            for row_dict in table.table_data:
                row_values = [row_dict.get(h, "") for h in headers]
                ws_table.append(row_values)
    try:
        wb.save(output_path)
        logger.info(f"Successfully exported data to '{output_path}'.")
    except Exception as e:
        logger.error(f"An unexpected error occurred during Excel export: {e}")
